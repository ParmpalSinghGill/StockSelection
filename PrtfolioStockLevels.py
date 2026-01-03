import pandas as pd
import DataLoad
import re,webbrowser
import openpyxl

stockfile="PortFolio/stock_analysis_export.xlsx"

# Load with openpyxl to get formulas
wb = openpyxl.load_workbook(stockfile)
ws = wb.active

data = []
headers = [cell.value for cell in ws[1]]
for row in ws.iter_rows(min_row=2, values_only=True):
    data.append(dict(zip(headers, row)))

stockdata = pd.DataFrame(data)
stockdata['Immediate Support'] = pd.to_numeric(stockdata['Immediate Support'], errors='coerce')

print(f"{'Stock':<40} {'Price':<10} {'Stoploss':<10} {'Diff %':<10} {'Status'} {'Screener Link'} {'TradingView Link'}")
print("-" * 90)


def extract_url(link_str):
    if pd.isna(link_str):
        return ""
    link_str = str(link_str)
    # Pattern to match =HYPERLINK("url", "text")
    match = re.search(r'=HYPERLINK\("([^"]+)"', link_str)
    if match:
        return match.group(1)
    return link_str

def printStocks():
    for index, row in stockdata.iterrows():
        stock_name = row['Stock Name']
        stoploss = row['Immediate Support']
        ScreenerLink = extract_url(row['Screener Link'])
        TradingView = extract_url(row['TradingView Link'])
        # print(row)

        # Skip if stoploss is not a number
        if pd.isna(stoploss) or not isinstance(stoploss, (int, float)):
            continue
            
        ticker = DataLoad.getTickerFromName(stock_name)
        if not ticker:
            # print(f"Ticker not found for {stock_name}")
            continue
            
        data = DataLoad.getData(ticker)
        if data is None or data.empty:
            # print(f"Data not found for {ticker}")
            continue
            
        latest_price = data['Close'].iloc[-1]
        
        # Check if price is near (within 3%) or below stoploss
        if latest_price <= stoploss * 1.03:
            diff_percent = ((latest_price - stoploss) / stoploss) * 100
            status = "BELOW" if latest_price < stoploss else "NEAR"
            
            print(f"{stock_name:<40} {latest_price:<10.2f} {stoploss:<10.2f} {diff_percent:<10.2f} {status} {ScreenerLink} {TradingView}")

def openTradingviewLinksInBrowser():
    for index, row in stockdata.iterrows():
        TradingView = extract_url(row['TradingView Link'])
        if TradingView:
            webbrowser.open(TradingView)



if __name__=="__main__":
    # printStocks()
    openTradingviewLinksInBrowser()
