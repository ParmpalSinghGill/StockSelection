import pandas as pd
import DataLoad
import re,webbrowser
import openpyxl



def readExcelFile(stockfile):
    # Load with openpyxl to get formulas
    wb = openpyxl.load_workbook(stockfile)
    ws = wb.active

    data = []
    headers = [cell.value for cell in ws[1]]
    for row in ws.iter_rows(min_row=2, values_only=True):
        data.append(dict(zip(headers, row)))

    stockdata = pd.DataFrame(data)
    stockdata['Immediate Support'] = pd.to_numeric(stockdata['Immediate Support'], errors='coerce')
    return stockdata

print(f"{'Stock':<40} {'Price':<10} {'Stoploss':<10} {'Diff %':<10} {'Status'} {'Screener Link'} {'TradingView Link'}")
print("-" * 110)

def calculate_smart_stoploss(support_price, buffer_pct=1.0):
    """
    Calculates a stop-loss price by subtracting a buffer percentage
    and rounding to the nearest valid tick size (0.05).
    """
    # 1. Apply the buffer (Subtract 0.5% to 1%)
    buffered_price = support_price * (1 - (buffer_pct / 100))
    
    # 2. Round to nearest 0.05 (Standard Tick Size)
    # This ensures you don't get weird numbers like 120.778
    final_stoploss = round(buffered_price / 0.05) * 0.05
    
    return round(final_stoploss, 2)

def extract_url(link_str):
    if pd.isna(link_str):
        return ""
    link_str = str(link_str)
    # Pattern to match =HYPERLINK("url", "text")
    match = re.search(r'=HYPERLINK\("([^"]+)"', link_str)
    if match:
        return match.group(1)
    return link_str

def printStocks(slpercentage=3):
    df=readExcelFile("PortFolio/Portfoliow_report.xlsx")
    df1=readExcelFile("PortFolio/Stocks_report.xlsx")
    # stockfile="PortFolio/Portfoliow_report.xlsx"
    stockdata=pd.concat([df,df1],axis=0)
    for index, row in stockdata.iterrows():
        stock_name = row['Stock Name']
        stoploss = calculate_smart_stoploss(row['Immediate Support'])
        ScreenerLink = extract_url(row['Screener Link'])
        TradingView = extract_url(row['TradingView Link'])

        # Skip if stoploss is not a number
        if pd.isna(stoploss) or not isinstance(stoploss, (int, float)):
            continue
            
        ticker = DataLoad.getTickerFromName(stock_name)
        if not ticker:
            # print(f"Ticker not found for {stock_name}")
            continue
            
        data = DataLoad.getData(ticker)
        if data is None or data.empty:
            # print(f"Data not found for {ticker}")
            continue
            
        latest_price = data['Close'].iloc[-1]
        
        # Check if price is near (within 3%) or below stoploss
        if latest_price <= stoploss * (1 + slpercentage/100):
            diff_percent = ((latest_price - stoploss) / stoploss) * 100
            status = "BELOW" if latest_price < stoploss else "NEAR"
            
            print(f"{stock_name:<40} {latest_price:<10.2f} {stoploss:<10.2f} {diff_percent:<10.2f} {status} {ScreenerLink} {TradingView}")

def openTradingviewLinksInBrowser(stockfile):
    stockdata=readExcelFile(stockfile)
    for index, row in stockdata.iterrows():
        TradingView = extract_url(row['TradingView Link'])
        if TradingView:
            webbrowser.open(TradingView)



if __name__=="__main__":
    printStocks(slpercentage=3)
    # openTradingviewLinksInBrowser("PortFolio/Portfoliow_report.xlsx")
